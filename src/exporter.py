"""
Exportador a Excel (.xlsx) usando openpyxl.
"""
import os
import json
import logging
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from src.models import ScrapeRun, QuoteResult
from src.config import DATA_DIR

logger = logging.getLogger(__name__)


def export_to_excel(scrape_run: ScrapeRun, filename: str = None) -> str:
    """
    Exporta los resultados de un ScrapeRun a un archivo Excel.
    Retorna la ruta del archivo generado.
    """
    if not filename:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"remesas_{ts}.xlsx"

    filepath = os.path.join(DATA_DIR, filename)
    os.makedirs(DATA_DIR, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Cotizaciones"

    # --- Styles ---
    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    number_format_clp = '#,##0'
    number_format_rate = '#,##0.0000'
    number_format_amount = '#,##0.00'

    # --- Headers ---
    headers = QuoteResult.csv_headers()
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # --- Data ---
    for row_idx, result in enumerate(scrape_run.results, 2):
        row_data = result.to_row()
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

            # Apply number formats
            if col in (6, 9, 10, 11):  # CLP amounts
                cell.number_format = number_format_clp
            elif col == 7:  # Monto recibido
                cell.number_format = number_format_amount
            elif col == 8:  # Tasa de cambio
                cell.number_format = number_format_rate

        # Alternate row colors
        if row_idx % 2 == 0:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = PatternFill(
                    start_color="F2F7FB", end_color="F2F7FB", fill_type="solid"
                )

    # --- Auto-width columns ---
    for col_idx in range(1, len(headers) + 1):
        max_width = len(str(headers[col_idx - 1]))
        for row_idx in range(2, min(len(scrape_run.results) + 2, 50)):
            cell_value = str(ws.cell(row=row_idx, column=col_idx).value or "")
            max_width = max(max_width, len(cell_value))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_width + 3, 30)

    # --- Freeze header row ---
    ws.freeze_panes = "A2"

    # --- Auto-filter ---
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(scrape_run.results) + 1}"

    # --- Metadata sheet ---
    ws_meta = wb.create_sheet("Metadata")
    ws_meta.cell(row=1, column=1, value="Timestamp").font = Font(bold=True)
    ws_meta.cell(row=1, column=2, value=scrape_run.timestamp)
    ws_meta.cell(row=2, column=1, value="Duración (seg)").font = Font(bold=True)
    ws_meta.cell(row=2, column=2, value=scrape_run.duration_seconds)
    ws_meta.cell(row=3, column=1, value="Total cotizaciones").font = Font(bold=True)
    ws_meta.cell(row=3, column=2, value=scrape_run.total_quotes)
    ws_meta.cell(row=4, column=1, value="Errores").font = Font(bold=True)
    ws_meta.cell(row=4, column=2, value=len(scrape_run.errors))
    for i, err in enumerate(scrape_run.errors):
        ws_meta.cell(row=5 + i, column=2, value=err)

    wb.save(filepath)
    logger.info(f"Excel exportado: {filepath}")
    return filepath


def save_json(scrape_run: ScrapeRun, filename: str = None) -> str:
    """Guarda los resultados como JSON."""
    if not filename:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"remesas_{ts}.json"

    filepath = os.path.join(DATA_DIR, filename)
    os.makedirs(DATA_DIR, exist_ok=True)

    with open(filepath, "w", encoding="utf-8") as f:
        f.write(scrape_run.to_json())

    logger.info(f"JSON guardado: {filepath}")
    return filepath


def load_latest_run() -> ScrapeRun | None:
    """Carga el último ScrapeRun desde data/."""
    if not os.path.exists(DATA_DIR):
        return None

    json_files = sorted(
        [f for f in os.listdir(DATA_DIR) if f.endswith(".json")],
        reverse=True
    )
    if not json_files:
        return None

    filepath = os.path.join(DATA_DIR, json_files[0])
    with open(filepath, "r", encoding="utf-8") as f:
        data = json.load(f)

    run = ScrapeRun(
        timestamp=data.get("timestamp", ""),
        duration_seconds=data.get("duration_seconds", 0),
        total_quotes=data.get("total_quotes", 0),
        errors=data.get("errors", []),
    )
    for r in data.get("results", []):
        run.results.append(QuoteResult(**r))

    return run
